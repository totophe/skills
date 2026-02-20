#!/usr/bin/env python3
"""Excel workbook reader for Claude Code. Streams large files efficiently."""

import subprocess
import sys

def ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("Installing openpyxl...", file=sys.stderr)
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
            stdout=sys.stderr, stderr=sys.stderr,
        )

ensure_openpyxl()

import argparse
import json
import os
from datetime import datetime, date, time

import openpyxl


def die(msg):
    print(f"Error: {msg}", file=sys.stderr)
    sys.exit(1)


def format_cell(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value)


def resolve_sheet(wb, name):
    if name is None:
        return wb[wb.sheetnames[0]]
    lower = name.lower()
    for sn in wb.sheetnames:
        if sn.lower() == lower:
            return wb[sn]
    available = ", ".join(wb.sheetnames)
    die(f"Sheet '{name}' not found. Available: {available}")


def open_workbook(path):
    if not os.path.isfile(path):
        die(f"File not found: {path}")
    if path.lower().endswith(".xls") and not path.lower().endswith(".xlsx"):
        die("Legacy .xls format not supported. Convert to .xlsx first.")
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        die(f"Cannot open file: {e}")


def parse_range(range_str):
    if range_str is None:
        return None, None
    if "-" in range_str:
        parts = range_str.split("-", 1)
        try:
            start = int(parts[0])
            end = int(parts[1])
        except ValueError:
            die(f"Invalid range '{range_str}'. Use N or N-M (e.g., 10, 10-20)")
        if start < 1 or end < start:
            die(f"Invalid range '{range_str}'. Start must be >= 1 and end >= start.")
        return start, end
    try:
        n = int(range_str)
        if n < 1:
            die(f"Row number must be >= 1, got {n}")
        return n, n
    except ValueError:
        die(f"Invalid range '{range_str}'. Use N or N-M (e.g., 10, 10-20)")


def read_header_row(ws, header_row):
    for i, row in enumerate(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), start=1):
        return [format_cell(c) for c in row]
    return []


def print_markdown_table(columns, rows, footer=None):
    col_widths = [len(str(c)) for c in columns]
    str_rows = []
    for row in rows:
        str_row = [str(v) for v in row]
        str_rows.append(str_row)
        for i, v in enumerate(str_row):
            if i < len(col_widths):
                col_widths[i] = max(col_widths[i], len(v))

    def fmt_row(vals):
        parts = []
        for i, v in enumerate(vals):
            w = col_widths[i] if i < len(col_widths) else len(v)
            parts.append(str(v).ljust(w))
        return "| " + " | ".join(parts) + " |"

    print(fmt_row(columns))
    print("|" + "|".join("-" * (w + 2) for w in col_widths) + "|")
    for row in str_rows:
        print(fmt_row(row))

    if footer:
        print(f"\n{footer}")


# --- Commands ---

def cmd_sheets(args):
    wb = open_workbook(args.file)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.max_row if ws.max_row else "unknown"
        cols = ws.max_column if ws.max_column else "unknown"
        sheets.append({"index": wb.sheetnames.index(name), "name": name, "rows": rows, "columns": cols})
    wb.close()

    if args.json:
        print(json.dumps({"file": args.file, "sheets": sheets}, indent=2, default=str))
    else:
        columns = ["#", "Sheet Name", "Rows", "Columns"]
        rows = [[s["index"] + 1, s["name"], s["rows"], s["columns"]] for s in sheets]
        print_markdown_table(columns, rows)


def cmd_headers(args):
    wb = open_workbook(args.file)
    ws = resolve_sheet(wb, args.sheet)
    num_rows = args.rows

    collected = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=num_rows, values_only=True), start=1):
        collected.append((i, [format_cell(c) for c in row]))
    wb.close()

    if not collected:
        die(f"Sheet '{ws.title}' is empty")

    if args.json:
        result = {
            "sheet": ws.title,
            "rows": [{"_row": r, "values": vals} for r, vals in collected],
        }
        print(json.dumps(result, indent=2))
    else:
        max_cols = max(len(vals) for _, vals in collected)
        col_headers = ["Row"] + [f"Col {i}" for i in range(1, max_cols + 1)]
        rows = [[r] + vals + [""] * (max_cols - len(vals)) for r, vals in collected]
        print_markdown_table(col_headers, rows, footer=f"Sheet: {ws.title}")


def cmd_rows(args):
    wb = open_workbook(args.file)
    ws = resolve_sheet(wb, args.sheet)
    header_row = args.header_row

    # Read headers
    headers = read_header_row(ws, header_row)
    if not headers:
        wb.close()
        die(f"Sheet '{ws.title}' is empty")

    # Determine row range
    range_start, range_end = parse_range(args.range)
    if range_start is not None:
        start_row = range_start
        limit = (range_end - range_start + 1) if range_end else args.limit
    else:
        start_row = header_row + 1 + args.offset
        limit = args.limit

    # Ensure we start after the header
    if start_row <= header_row:
        start_row = header_row + 1

    collected = []
    count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if range_end is not None and start_row + count > range_end:
            break
        collected.append((start_row + count, [format_cell(c) for c in row]))
        count += 1
        if count >= limit:
            break

    total = ws.max_row
    wb.close()

    if not collected:
        print("No data rows found.")
        return

    if args.json:
        result = {
            "sheet": ws.title,
            "headers": headers,
            "rows": [],
            "total_rows": total,
            "showing": len(collected),
        }
        for row_num, vals in collected:
            obj = {"_row": row_num}
            for i, h in enumerate(headers):
                obj[h if h else f"col_{i}"] = vals[i] if i < len(vals) else ""
            result["rows"].append(obj)
        print(json.dumps(result, indent=2))
    else:
        col_headers = ["Row"] + headers
        rows = [[r] + vals + [""] * (len(headers) - len(vals)) for r, vals in collected]
        total_str = f"{total:,}" if isinstance(total, int) else "unknown"
        footer = f"Showing {len(collected)} rows (of {total_str} total). Sheet: {ws.title}"
        print_markdown_table(col_headers, rows, footer=footer)


def cmd_column(args):
    wb = open_workbook(args.file)
    ws = resolve_sheet(wb, args.sheet)
    header_row = args.header_row

    # Read headers
    headers = read_header_row(ws, header_row)
    if not headers:
        wb.close()
        die(f"Sheet '{ws.title}' is empty")

    # Resolve column
    col_idx = None
    col_name = args.column

    # Try case-insensitive name match
    lower_headers = [h.lower() for h in headers]
    if col_name.lower() in lower_headers:
        col_idx = lower_headers.index(col_name.lower())
        col_name = headers[col_idx]
    else:
        # Try as integer index
        try:
            idx = int(col_name)
            if 0 <= idx < len(headers):
                col_idx = idx
                col_name = headers[col_idx]
            else:
                die(f"Column index {idx} out of range (0-{len(headers) - 1}). Available: {', '.join(headers)}")
        except ValueError:
            die(f"Column '{col_name}' not found. Available: {', '.join(headers)}")

    # Stream column data
    start_row = header_row + 1 + args.offset
    collected = []
    count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        val = format_cell(row[col_idx]) if col_idx < len(row) else ""
        collected.append((start_row + count, val))
        count += 1
        if count >= args.limit:
            break

    total = ws.max_row
    wb.close()

    if not collected:
        print("No data rows found.")
        return

    if args.json:
        result = {
            "sheet": ws.title,
            "column": col_name,
            "column_index": col_idx,
            "values": [{"_row": r, "value": v} for r, v in collected],
            "total_rows": total,
            "showing": len(collected),
        }
        print(json.dumps(result, indent=2))
    else:
        col_headers = ["Row", col_name]
        rows = [[r, v] for r, v in collected]
        total_str = f"{total:,}" if isinstance(total, int) else "unknown"
        footer = f"Column \"{col_name}\" (index {col_idx}). Showing {len(collected)} rows (of {total_str} total). Sheet: {ws.title}"
        print_markdown_table(col_headers, rows, footer=footer)


def build_parser():
    parser = argparse.ArgumentParser(description="Excel workbook reader")
    sub = parser.add_subparsers(dest="command", required=True)

    # sheets
    p = sub.add_parser("sheets", help="List all sheets with dimensions")
    p.add_argument("file", help="Path to Excel file")
    p.add_argument("--json", action="store_true", help="Output as JSON")

    # headers
    p = sub.add_parser("headers", help="Show header rows")
    p.add_argument("file", help="Path to Excel file")
    p.add_argument("--sheet", default=None, help="Sheet name (case-insensitive)")
    p.add_argument("--rows", type=int, default=1, help="Number of rows to show (default: 1)")
    p.add_argument("--json", action="store_true", help="Output as JSON")

    # rows
    p = sub.add_parser("rows", help="Extract rows")
    p.add_argument("file", help="Path to Excel file")
    p.add_argument("range", nargs="?", default=None, help="Row range: N or N-M (1-based)")
    p.add_argument("--sheet", default=None, help="Sheet name (case-insensitive)")
    p.add_argument("--limit", type=int, default=50, help="Max rows to return (default: 50)")
    p.add_argument("--offset", type=int, default=0, help="Skip N data rows")
    p.add_argument("--header-row", type=int, default=1, help="Row containing headers (default: 1)")
    p.add_argument("--json", action="store_true", help="Output as JSON")

    # column
    p = sub.add_parser("column", help="Extract a column")
    p.add_argument("file", help="Path to Excel file")
    p.add_argument("column", help="Column name (case-insensitive) or 0-based index")
    p.add_argument("--sheet", default=None, help="Sheet name (case-insensitive)")
    p.add_argument("--limit", type=int, default=50, help="Max rows to return (default: 50)")
    p.add_argument("--offset", type=int, default=0, help="Skip N data rows")
    p.add_argument("--header-row", type=int, default=1, help="Row containing headers (default: 1)")
    p.add_argument("--json", action="store_true", help="Output as JSON")

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    commands = {
        "sheets": cmd_sheets,
        "headers": cmd_headers,
        "rows": cmd_rows,
        "column": cmd_column,
    }

    try:
        commands[args.command](args)
    except KeyboardInterrupt:
        sys.exit(130)
    except Exception as e:
        die(str(e))


if __name__ == "__main__":
    main()
